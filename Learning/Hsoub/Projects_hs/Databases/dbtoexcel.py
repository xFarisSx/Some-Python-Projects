import sqlite3
from openpyxl import Workbook

con = sqlite3.connect('database.db')
cur = con.cursor()

cur.execute("SELECT * FROM employees")
# print(cur.fetchall())
# print(cur.fetchone())
# print(cur.fetchmany(3))

answer = cur.fetchall()
for i in answer:
    print(i)

excel = Workbook()
sheet = excel.create_sheet('main', 0)

for y in range(len(answer)):
    for x in range(len(answer[y])):
        sheet.cell(row=y+1, column=x+1).value = answer[y][x]
excel.save("dbtoexcel.xlsx")