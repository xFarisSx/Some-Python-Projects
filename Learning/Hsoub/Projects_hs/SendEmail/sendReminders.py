import openpyxl, smtplib


sender_email = 'faristaleek@gmail.com'
password = 'vmxxwkbsikrflxsz'
server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(sender_email, password)

file = openpyxl.load_workbook('members.xlsx')
sheet = file.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value
print(latestMonth)

unpaidMembers= {}
for r in range(2, sheet.max_row+1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email



for name, email in unpaidMembers.items():
    body = """Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s.Please make this payment as soon as possible.Thanks!""" %(latestMonth, name, latestMonth)
    sendMailStatus = server.sendmail(sender_email, email, body)

    if sendMailStatus != {} :
        print(f'there is a problem, {email}, {sendMailStatus}')


server.quit()
