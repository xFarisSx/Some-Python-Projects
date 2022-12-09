import time
import openpyxl
from selenium import webdriver
import chromedriver_binary
from selenium.webdriver.common.by import By
import csv
from pathlib import Path

browser = webdriver.Chrome()
browser.get('https://en.wikipedia.org/wiki/List_of_languages_by_number_of_native_speakers')

tables = browser.find_elements(By.TAG_NAME, 'table')
requiredTable = tables[1]
print(requiredTable)

rows = requiredTable.find_elements(By.TAG_NAME, 'tr')
print(len(rows))
headers = []
for header in rows[0].find_elements(By.TAG_NAME, 'th'):
    headers.append(header.text.strip().replace('\n', ' '))
print(headers)

data = []

for rowData in rows:
    value = rowData.find_elements(By.TAG_NAME, 'td')

    value_text = [db.text.strip() for db in value]

    if len(value_text) == 0: continue

    data.append(value_text)
print(data)

excelFile = openpyxl.Workbook()
excelFile.create_sheet(index=0, title='table')

sheet = excelFile['table']

for i in range(len(headers)):
    sheet.cell(row=1, column=i+1).value = headers[i]

for row in range(len(data)):
    for col in range(len(data[row])):
        sheet.cell(row=row+2, column=col+1).value = data[row][col]

excelFile.save(filename='table.xlsx')