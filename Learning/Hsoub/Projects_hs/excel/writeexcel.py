import openpyxl

#  create file
excelFile = openpyxl.Workbook()
print(excelFile.sheetnames)

# chane name
excelSheet = excelFile.active
excelSheet.title = 'firstSheet'

# create sheet
excelFile.create_sheet()
excelFile.create_sheet()
excelFile.create_sheet(index=1, title='secondSheet')

# delete sheet
del excelFile['Sheet1']

# write to sheet
sheet = excelFile['secondSheet']
sheet['A1'] = 'Hello World'
print(sheet['A1'].value)

# practice
names = ['hadi','faris','omar','anas']
firstSheet = excelFile['firstSheet']
for i in range(1,len(names)+1):
    firstSheet.cell(row=i, column=3).value = names[i-1]

#  save file
excelFile.save(filename='newExcel.xlsx')