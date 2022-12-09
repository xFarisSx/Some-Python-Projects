import openpyxl

# open excel file
excelFile = openpyxl.load_workbook('example.xlsx')
# print sheetnames
print(excelFile.sheetnames)

sheet1 = excelFile['Sheet1']

# sheet name
print(sheet1.title)

# print the opened sheet
activeSheet = excelFile.active
print(activeSheet.title)

# cell value
print(sheet1['A1'].value)
print(sheet1['B1'].value)
print(sheet1['C1'].value)

# print row and column
print(sheet1['C1'].row)
print(sheet1['C1'].column)

# coordinate
print(sheet1['C1'].coordinate)

# print cell value
print(sheet1.cell(row=1,column=2).value)

for i in range(1,7):
    print(i, sheet1.cell(row=i,column=3).value)

print('-'*50)

total = 0
for i in range(1,sheet1.max_row+1):
    print(i, sheet1.cell(row=i,column=1).value, sheet1.cell(row=i,column=2).value)
    total += sheet1.cell(row=i,column=2).value
print("The total salary is",total)

print("-"*50)
# row and column number
print(sheet1.max_row)
print(sheet1.max_column)











